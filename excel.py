import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# API endpoints
FINANCE_URL = "https://api.jkuatcu.org/finance"
FINANCE_DATA_URL = "https://api.jkuatcu.org/financedata"
HEADERS = {"Origin": "https://admin.jkuatcu.org"}

def fetch_data(url):
    """Fetch data from the given URL with error handling."""
    try:
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()
        return response.json() or {}
    except requests.RequestException as e:
        logging.error(f"Error fetching data from {url}: {e}")
        return {}

def safe_float(value, default=0.0):
    """Safely convert a value to float, returning a default if conversion fails."""
    try:
        return float(value) if value is not None else default
    except (ValueError, TypeError):
        return default

def get_approved_budget(department_id, semester, approved_data):
    """Get the approved budget for a department and semester."""
    for budget in approved_data.get('budgets', []) or []:
        if str(budget.get('department_id', '')) == str(department_id) and budget.get('semester') == semester:
            return budget
    return {}

def generate_excel(budgets, approved_data):
    """Generate an Excel file with grouped budget data including assets and approved columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Data"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    row = 1
    for semester in sorted(set(b.get('semester', 'N/A') for b in budgets)):
        ws.cell(row=row, column=1, value=f"Semester: {semester}").font = Font(bold=True)
        row += 2

        for budget in sorted((b for b in budgets if b.get('semester', 'N/A') == semester), key=lambda x: x.get('department_name', '')):
            department_name = str(budget.get('department_name', 'N/A'))
            department_id = budget.get('department_id')
            ws.cell(row=row, column=2, value=f"Department: {department_name}").font = Font(bold=True)
            row += 2

            approved_budget = get_approved_budget(department_id, semester, approved_data)
            
            for event in budget.get('events', []) or []:
                ws.cell(row=row, column=2, value=f"Event: {event.get('name', 'N/A')}").font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=2, value=f"Attendance: {event.get('attendance', 'N/A')}")
                row += 1
                
                headers = ["Item", "Price", "Quantity", "Total", "Approved"]
                for col, header in enumerate(headers, start=2):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.border = thick_border
                row += 1
                
                for item in event.get('items', []) or []:
                    quantity = safe_float(item.get('quantity', 0))
                    price = safe_float(item.get('price', 0))
                    total = quantity * price
                    
                    ws.cell(row=row, column=2, value=item.get('name', 'N/A')).border = border
                    ws.cell(row=row, column=3, value=price).border = border
                    ws.cell(row=row, column=4, value=quantity).border = border
                    ws.cell(row=row, column=5, value=total).border = border
                    
                    approved_item = next(
                        (a for e in (approved_budget.get('events') or []) if isinstance(e, dict)
                         for a in (e.get('items') or []) if isinstance(a, dict) and a.get('name') == item.get('name')),
                        None
                    )
                    approved_total = safe_float(approved_item.get('quantity', 0)) * safe_float(approved_item.get('price', 0)) if approved_item else 0
                    ws.cell(row=row, column=6, value=approved_total).border = border
                    row += 1
                row += 1
            
            ws.cell(row=row, column=2, value="Assets").font = Font(bold=True)
            row += 1
            
            for asset in budget.get('assets', []) or []:
                quantity = safe_float(asset.get('quantity', 0))
                price = safe_float(asset.get('price', 0))
                total = quantity * price
                
                ws.cell(row=row, column=2, value=asset.get('name', 'N/A')).border = border
                ws.cell(row=row, column=3, value=price).border = border
                ws.cell(row=row, column=4, value=quantity).border = border
                ws.cell(row=row, column=5, value=total).border = border
                
                approved_asset = next((a for a in (approved_budget.get('assets') or []) if isinstance(a, dict) and a.get('name') == asset.get('name')), None)
                approved_total = safe_float(approved_asset.get('quantity', 0)) * safe_float(approved_asset.get('price', 0)) if approved_asset else 0
                ws.cell(row=row, column=6, value=approved_total).border = border
                row += 1
            row += 2  
    
    filename = f"detailed_budget_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    logging.info(f"Excel file '{filename}' has been created successfully.")

def main():
    logging.info("Fetching budget data...")
    budgets_data = fetch_data(FINANCE_URL)
    logging.info("Fetching approved budget data...")
    approved_data = fetch_data(FINANCE_DATA_URL)
    logging.info("Generating Excel file...")
    generate_excel(budgets_data.get('budgets', []), approved_data)

if __name__ == "__main__":
    main()
